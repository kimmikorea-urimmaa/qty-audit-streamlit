#!/usr/bin/env python3
"""조경 시설물 수량산출서 자동검토 도구 (오탐 저감 버전)

핵심 정책
- 할증 관련 검토(allowance_policy_check / allowance_check)는 "비고(BIGO)에 %가 있는 행"에서만 수행한다.
- calc_text_check / unit_weight는 비고% 유무와 무관하게 항상 수행한다.

오탐을 줄이기 위한 핵심 개선
1) E 수식이 =ROUND(셀*계수, n) 형태면, 참조셀의 "값(data_only)"을 가져와 엑셀 방식대로 expected를 계산한다.
   (단계별 ROUND 때문에 0.001 차이가 나는 케이스를 정상 처리)
2) allowance_check에서 D 산출근거에 이미 *1.04 같은 할증계수가 포함되어 있으면 "또 곱하지 않음".
"""

from __future__ import annotations

import argparse
import ast
import csv
import math
import os
import re
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple


@dataclass
class ErrorRecord:
    row: int
    cell: str
    check_type: str
    reason: str
    severity: str
    related_formula: str = ""
    actual_value: Optional[float] = None
    expected_value: Optional[float] = None
    difference: Optional[float] = None
    tol: Optional[float] = None
    rule_name: str = ""


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="시설물 수량산출서 자동검토")
    p.add_argument("xlsx", help="입력 XLSX 파일 경로")
    p.add_argument("--rules", default="rules.yml", help="룰 YAML 파일 경로")
    p.add_argument("--outdir", default="output", help="결과 출력 폴더")
    return p.parse_args()


def load_rules(path: str) -> Dict[str, Any]:
    try:
        import yaml
    except Exception as exc:
        raise RuntimeError("pyyaml 미설치: `pip install pyyaml` 필요") from exc

    with open(path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f) or {}


def normalize_text(v: Any) -> str:
    return "" if v is None else str(v).strip()


def choose_sheet_name(wb) -> str:
    if "시설물산출" in wb.sheetnames:
        return "시설물산출"

    scored: List[Tuple[int, str]] = []
    for name in wb.sheetnames:
        score = 0
        if "시설물" in name:
            score += 2
        if "산출" in name:
            score += 2
        if "수량" in name:
            score += 1
        if score > 0:
            scored.append((score, name))

    if scored:
        scored.sort(reverse=True)
        return scored[0][1]

    return wb.sheetnames[0]


def detect_columns(ws) -> Tuple[int, Dict[str, int]]:
    """헤더에서 열 자동 탐지, 실패 시 기본 B~G 사용."""
    columns = {"work": 2, "spec": 3, "basis": 4, "qty": 5, "unit": 6, "bigo": 7}
    header_row = 1

    for r in range(1, min(ws.max_row, 20) + 1):
        row_values = [normalize_text(ws.cell(r, c).value) for c in range(1, min(ws.max_column, 40) + 1)]
        hit = 0
        for idx, text in enumerate(row_values, start=1):
            low = text.lower()
            if "공종" in text:
                columns["work"] = idx
                hit += 1
            if "규격" in text:
                columns["spec"] = idx
                hit += 1
            if "산출근거" in text:
                columns["basis"] = idx
                hit += 1
            if "수량" in text:
                columns["qty"] = idx
                hit += 1
            if "단위" in text:
                columns["unit"] = idx
                hit += 1
            if "비고" in text or "remark" in low:
                columns["bigo"] = idx
                hit += 1
        if hit >= 2:
            header_row = r
            break

    return header_row, columns


def as_float(v: Any) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
            return None
        return float(v)
    s = str(v).strip().replace(",", "")
    try:
        return float(s)
    except Exception:
        return None


def has_cell_reference(expr: str) -> bool:
    return bool(re.search(r"\$?[A-Za-z]{1,3}\$?\d+", expr))


def safe_eval_numeric(expr: str) -> Optional[float]:
    """숫자/연산자/괄호만 허용해 계산."""
    expr = expr.strip()
    if expr.startswith("="):
        expr = expr[1:].strip()

    allowed_nodes = (
        ast.Expression,
        ast.BinOp,
        ast.UnaryOp,
        ast.Constant,
        ast.Add,
        ast.Sub,
        ast.Mult,
        ast.Div,
        ast.Pow,
        ast.USub,
        ast.UAdd,
        ast.Mod,
        ast.FloorDiv,
    )

    try:
        tree = ast.parse(expr, mode="eval")
        for node in ast.walk(tree):
            if not isinstance(node, allowed_nodes):
                return None
        result = eval(compile(tree, "<expr>", "eval"), {"__builtins__": {}}, {})
        return as_float(result)
    except Exception:
        return None


def parse_round_digits(formula: str) -> Optional[int]:
    if not formula:
        return None
    m = re.search(r"ROUND\s*\(.*?,\s*(-?\d+)\s*\)", formula, flags=re.IGNORECASE)
    if not m:
        return None
    try:
        return int(m.group(1))
    except Exception:
        return None


def get_round_digits_for_row(e_formula: str, default_digits: int) -> int:
    n = parse_round_digits(e_formula)
    return n if n is not None else default_digits


def tol_from_round_digits(round_digits: int) -> float:
    # 3자리 반올림이면 0.0005가 이론상 최대오차.
    # float/엑셀 차이 흡수 위해 약간 넉넉히 0.0006.
    return 0.6 * (10 ** (-round_digits))


def classify_row_type(work: str, spec: str, unit: str, bigo: str, rules: Dict[str, Any]) -> str:
    """material / installation / unknown (키워드 기반)"""
    text = f"{work} {spec} {unit} {bigo}".lower()
    material_keys = [str(x).lower() for x in (rules.get("material_keywords_any") or [])]
    install_keys = [str(x).lower() for x in (rules.get("installation_keywords_any") or [])]

    if material_keys and any(k in text for k in material_keys):
        return "material"
    if install_keys and any(k in text for k in install_keys):
        return "installation"
    return "unknown"


def unit_weight_check(work: str, spec: str, unit: str) -> List[Tuple[str, str, str]]:
    issues: List[Tuple[str, str, str]] = []
    w = f"{work} {spec}".lower()
    u = unit.strip().lower()

    if "아연도각관" in f"{work} {spec}":
        if u == "kg":
            issues.append(("HIGH", "아연도각관은 m 단가 처리 가능 품목인데 kg 단위로 입력됨", "unit_weight:아연도각관"))
        return issues

    if "st pl" in w or "sts pl" in w:
        if not re.search(r"\bT\s*\d+(\.\d+)?\b", spec, flags=re.IGNORECASE):
            issues.append(("MEDIUM", "PL 품목인데 규격에 두께(T값) 정보가 없음", "unit_weight:plate-thickness"))

    if "angle" in w and u in {"m", "m2", "㎡"}:
        issues.append(("LOW", "angle 품목은 39.65 kg/m2 기준 검토 대상", "unit_weight:angle-39.65"))

    if "이형철근" in w:
        if not re.search(r"\bD\s*\d+\b", spec, flags=re.IGNORECASE):
            issues.append(("MEDIUM", "이형철근 품목인데 규격에 D값이 없음", "unit_weight:rebar-diameter"))

    return issues


def eval_e_round_ref_formula(e_formula: str, ws_value) -> Optional[Tuple[float, str, float, int]]:
    """
    지원: =ROUND(E146*1.04,3) / =ROUND(E146,3)
    반환: (expected_value, ref_cell, multiplier_in_formula, digits)
    """
    if not e_formula:
        return None

    f = e_formula.strip()
    if f.startswith("="):
        f = f[1:].strip()

    m = re.match(
        r"ROUND\s*\(\s*([A-Za-z]{1,3}\d+)\s*(?:\*\s*([0-9]+(?:\.[0-9]+)?))?\s*,\s*(-?\d+)\s*\)\s*$",
        f,
        flags=re.IGNORECASE,
    )
    if not m:
        return None

    ref_cell = m.group(1).upper()
    mult = float(m.group(2)) if m.group(2) else 1.0
    digits = int(m.group(3))

    v = as_float(ws_value[ref_cell].value)
    if v is None:
        return None

    expected = round(v * mult, digits)
    return expected, ref_cell, mult, digits


def eval_e_round_pure_formula(e_formula: str) -> Optional[Tuple[float, int]]:
    """
    지원: =ROUND(0.7*0.13*2.5,3) 처럼 셀참조 없는 ROUND(표현식,n)
    반환: (expected_value, digits)
    """
    if not e_formula:
        return None
    f = e_formula.strip()
    if f.startswith("="):
        f = f[1:].strip()

    m = re.match(r"ROUND\s*\(\s*(.+)\s*,\s*(-?\d+)\s*\)\s*$", f, flags=re.IGNORECASE)
    if not m:
        return None

    inner = m.group(1).strip()
    digits = int(m.group(2))

    # inner에 셀참조가 있으면 여기서 계산하지 않음
    if has_cell_reference(inner):
        return None

    val = safe_eval_numeric(inner)
    if val is None:
        return None

    return round(val, digits), digits


def d_has_multiplier(d_text: str, mult: float) -> bool:
    """
    D 산출근거에 '* 1.04' 같이 계수가 직접 포함되어 있는지 대략 감지.
    (이중 할증 방지)
    """
    if not d_text:
        return False
    # 곱해지는 숫자들 중 mult에 가까운 값이 있는지 확인
    candidates = re.findall(r"\*\s*([0-9]+(?:\.[0-9]+)?)", d_text.replace(",", ""))
    for s in candidates:
        try:
            if abs(float(s) - mult) < 1e-9:
                return True
        except Exception:
            pass
    return False


def build_reports(errors: List[ErrorRecord], outdir: str) -> None:
    from openpyxl import Workbook

    os.makedirs(outdir, exist_ok=True)
    csv_path = os.path.join(outdir, "report.csv")
    xlsx_path = os.path.join(outdir, "report.xlsx")

    columns = [
        "row",
        "cell",
        "check_type",
        "reason",
        "severity",
        "rule_name",
        "related_formula",
        "actual_value",
        "expected_value",
        "difference",
        "tol",
    ]

    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(columns)
        for e in errors:
            w.writerow(
                [
                    e.row,
                    e.cell,
                    e.check_type,
                    e.reason,
                    e.severity,
                    e.rule_name,
                    e.related_formula,
                    e.actual_value,
                    e.expected_value,
                    e.difference,
                    e.tol,
                ]
            )

    summary: Dict[Tuple[str, str], int] = {}
    for e in errors:
        summary[(e.check_type, e.severity)] = summary.get((e.check_type, e.severity), 0) + 1

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(["check_type", "severity", "count"])
    for (ct, sev), cnt in sorted(summary.items(), key=lambda x: (x[0][0], x[0][1])):
        ws_summary.append([ct, sev, cnt])

    ws_errors = wb.create_sheet("Errors")
    ws_errors.append(columns)
    for e in errors:
        ws_errors.append(
            [
                e.row,
                e.cell,
                e.check_type,
                e.reason,
                e.severity,
                e.rule_name,
                e.related_formula,
                e.actual_value,
                e.expected_value,
                e.difference,
                e.tol,
            ]
        )

    wb.save(xlsx_path)


def main() -> None:
    args = parse_args()

    try:
        rules = load_rules(args.rules)
    except Exception as e:
        raise SystemExit(f"[ERROR] rules 로드 실패: {e}")

    try:
        from openpyxl import load_workbook

        wb_formula = load_workbook(args.xlsx, data_only=False)
        wb_value = load_workbook(args.xlsx, data_only=True)
    except Exception as e:
        raise SystemExit(f"[ERROR] xlsx 로드 실패: {e}")

    sheet_name = choose_sheet_name(wb_formula)
    ws_formula = wb_formula[sheet_name]
    ws_value = wb_value[sheet_name]

    header_row, cols = detect_columns(ws_formula)

    percent_regex = re.compile(rules.get("allowance_percent_extract_regex", r"(\d+(\.\d+)?)%"))
    allowance_map = rules.get("allowance_multiplier_map", {})

    default_round_digits = int(rules.get("round_default_digits", 3))
    sev_install_has_allowance = str(rules.get("policy_installation_has_allowance_severity", "HIGH"))

    errors: List[ErrorRecord] = []

    for r in range(header_row + 1, ws_formula.max_row + 1):
        work = normalize_text(ws_formula.cell(r, cols["work"]).value)
        spec = normalize_text(ws_formula.cell(r, cols["spec"]).value)
        d_text = normalize_text(ws_formula.cell(r, cols["basis"]).value)
        e_formula = normalize_text(ws_formula.cell(r, cols["qty"]).value)
        e_value = as_float(ws_value.cell(r, cols["qty"]).value)
        unit = normalize_text(ws_formula.cell(r, cols["unit"]).value)
        bigo = normalize_text(ws_formula.cell(r, cols["bigo"]).value)

        if not any([work, spec, d_text, e_formula, unit, bigo]):
            continue

        # 행 유형
        row_type = classify_row_type(work, spec, unit, bigo, rules)

        # ROUND 자리수/허용오차
        round_digits = get_round_digits_for_row(e_formula, default_round_digits)
        tol = tol_from_round_digits(round_digits)

        # -----------------------------
        # (1) calc_text_check (항상)
        # -----------------------------
        expected_from_e: Optional[float] = None
        expected_rule_name = ""

        # 1-A) E가 ROUND(셀*계수,n) 형태면, 참조셀 값 기반으로 expected 계산(단계별 ROUND 대응)
        ref_eval = eval_e_round_ref_formula(e_formula, ws_value)
        if ref_eval is not None:
            expected_from_e, ref_cell, mult_in_e, digits_in_e = ref_eval
            # E 수식의 digits가 실제 기준
            round_digits = digits_in_e
            tol = tol_from_round_digits(round_digits)
            expected_rule_name = f"E:ROUND({ref_cell}*{mult_in_e},{digits_in_e})"

        # 1-B) E가 ROUND(순수식,n) 형태면 계산
        if expected_from_e is None:
            pure_eval = eval_e_round_pure_formula(e_formula)
            if pure_eval is not None:
                expected_from_e, digits_in_e = pure_eval
                round_digits = digits_in_e
                tol = tol_from_round_digits(round_digits)
                expected_rule_name = f"E:ROUND(pure_expr,{digits_in_e})"

        # 1-C) 위 2개가 아니면 D(산출근거) 계산값을 ROUND해서 비교(기존 방식)
        d_numeric: Optional[float] = None
        if expected_from_e is None:
            if d_text and not has_cell_reference(d_text):
                d_numeric = safe_eval_numeric(d_text)
                if d_numeric is not None:
                    expected_from_e = round(d_numeric, round_digits)
                    expected_rule_name = f"D_round({round_digits})"

        if expected_from_e is not None and e_value is not None:
            diff = abs(expected_from_e - e_value)
            if diff > tol:
                errors.append(
                    ErrorRecord(
                        row=r,
                        cell=f"D{r}/E{r}",
                        check_type="calc_text_check",
                        reason=f"계산 기대값(ROUND {round_digits})과 E 수량 불일치",
                        severity="HIGH",
                        related_formula=f"D:{d_text} | E:{e_formula} | BIGO:{bigo}",
                        actual_value=e_value,
                        expected_value=expected_from_e,
                        difference=diff,
                        tol=tol,
                        rule_name=expected_rule_name or f"ROUND({round_digits}) 비교",
                    )
                )

        # -----------------------------
        # (2) unit_weight (항상)
        # -----------------------------
        for sev, reason, rule_name in unit_weight_check(work, spec, unit):
            errors.append(
                ErrorRecord(
                    row=r,
                    cell=f"F{r}",
                    check_type="unit_weight",
                    reason=reason,
                    severity=sev,
                    rule_name=rule_name,
                )
            )

        # -----------------------------
        # (3) 할증 검토 (비고% 있을 때만)
        # -----------------------------
        m = percent_regex.search(bigo)
        if not m:
            continue  # 할증 관련만 스킵

        percent_text = f"{m.group(1)}%"
        if percent_text not in allowance_map:
            errors.append(
                ErrorRecord(
                    row=r,
                    cell=f"G{r}",
                    check_type="allowance_check",
                    reason=f"비고에 '{percent_text}'가 있으나 allowance_multiplier_map에 정의되지 않음",
                    severity="MEDIUM",
                    related_formula=f"BIGO:{bigo} | E:{e_formula}",
                    rule_name="allowance_multiplier_map missing",
                )
            )
            continue

        multiplier = float(allowance_map[percent_text])
        rule_name = f"비고 퍼센트({percent_text})"

        # 3-A) 설치품인데 비고에 %가 있으면 정책 위반
        if row_type == "installation":
            errors.append(
                ErrorRecord(
                    row=r,
                    cell=f"E{r}",
                    check_type="allowance_policy_check",
                    reason="설치품(정미량) 항목인데 비고에 할증(%)이 명시됨",
                    severity=sev_install_has_allowance,
                    related_formula=f"D:{d_text} | E:{e_formula} | BIGO:{bigo}",
                    rule_name=rule_name,
                )
            )
            continue

        # 3-B) 재료에서만 allowance_check 수행
        if row_type == "material" and e_value is not None:
            # 3-B-1) E가 ROUND(셀*계수,n)라면 그 계수(수식상 계수)가 비고%와 맞는지 + 결과가 맞는지 검증
            ref_eval2 = eval_e_round_ref_formula(e_formula, ws_value)
            if ref_eval2 is not None:
                expected_e, ref_cell, mult_in_e, digits_in_e = ref_eval2
                tol2 = tol_from_round_digits(digits_in_e)

                # (a) 계수 자체가 비고%와 다르면 잡아줌(원하면 끌 수도 있음)
                if abs(mult_in_e - multiplier) > 1e-9:
                    errors.append(
                        ErrorRecord(
                            row=r,
                            cell=f"E{r}",
                            check_type="allowance_check",
                            reason=f"비고 할증({multiplier})과 E 수식 계수({mult_in_e})가 다름",
                            severity="MEDIUM",
                            related_formula=f"E:{e_formula} | BIGO:{bigo}",
                            actual_value=mult_in_e,
                            expected_value=multiplier,
                            difference=abs(mult_in_e - multiplier),
                            tol=0.0,
                            rule_name=rule_name,
                        )
                    )

                # (b) 계산 결과 비교
                diff2 = abs(expected_e - e_value)
                if diff2 > tol2:
                    errors.append(
                        ErrorRecord(
                            row=r,
                            cell=f"E{r}",
                            check_type="allowance_check",
                            reason=f"재료 항목: E 수식(참조셀 기반) 계산값과 E 수량 불일치",
                            severity="MEDIUM",
                            related_formula=f"E:{e_formula} | BIGO:{bigo}",
                            actual_value=e_value,
                            expected_value=expected_e,
                            difference=diff2,
                            tol=tol2,
                            rule_name=f"{rule_name} | E참조({ref_cell})",
                        )
                    )
                continue

            # 3-B-2) E가 참조형이 아니면, D 수식으로 검증(단, D에 이미 계수가 있으면 이중 곱 금지)
            if d_numeric is None and d_text and not has_cell_reference(d_text):
                d_numeric = safe_eval_numeric(d_text)

            if d_numeric is not None:
                # D에 이미 *1.04 같은 계수가 있으면 expected = D(그 자체)로 비교
                if d_has_multiplier(d_text, multiplier):
                    expected_allow = round(d_numeric, round_digits)
                    rule2 = f"{rule_name} | D already has {multiplier}"
                else:
                    expected_allow = round(d_numeric * multiplier, round_digits)
                    rule2 = f"{rule_name} | D*{multiplier}"

                diff3 = abs(expected_allow - e_value)
                tol3 = tol_from_round_digits(round_digits)
                if diff3 > tol3:
                    errors.append(
                        ErrorRecord(
                            row=r,
                            cell=f"E{r}",
                            check_type="allowance_check",
                            reason=f"재료 항목: 비고 할증 적용 기대값과 E 수량 불일치",
                            severity="MEDIUM",
                            related_formula=f"D:{d_text} | E:{e_formula} | BIGO:{bigo}",
                            actual_value=e_value,
                            expected_value=expected_allow,
                            difference=diff3,
                            tol=tol3,
                            rule_name=rule2,
                        )
                    )

    try:
        build_reports(errors, args.outdir)
    except Exception as e:
        raise SystemExit(f"[ERROR] 리포트 저장 실패: {e}")

    print(f"[OK] sheet='{sheet_name}', header_row={header_row}, rows_checked={ws_formula.max_row - header_row}")
    print(f"[OK] reports saved to: {os.path.join(args.outdir, 'report.csv')} / report.xlsx")


if __name__ == "__main__":
    main()
